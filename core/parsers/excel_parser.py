import openpyxl


ROWS_PER_PAGE = 100


def parse_excel(file_path: str) -> list[str]:
    """Parse an Excel file. Each sheet is split into pages of ~100 rows.
    Handles Bloomberg-style multi-header formats.
    Returns list of strings — one per page.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    pages = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_str = "\t".join(
                str(cell) if cell is not None else "" for cell in row
            )
            if row_str.strip():
                rows.append(row_str)

        if not rows:
            continue

        # Split into pages of ROWS_PER_PAGE rows
        # Always include first 5 rows (headers) in each page for context
        header_rows = rows[:5]
        data_rows = rows[5:]

        if not data_rows:
            # Sheet has only headers
            page_text = f"[Sheet: {sheet_name}]\n" + "\n".join(rows)
            pages.append(page_text)
            continue

        for i in range(0, len(data_rows), ROWS_PER_PAGE):
            batch = data_rows[i : i + ROWS_PER_PAGE]
            page_text = (
                f"[Sheet: {sheet_name}, Rows {i + 6}-{i + 5 + len(batch)}]\n"
                + "\n".join(header_rows)
                + "\n"
                + "\n".join(batch)
            )
            pages.append(page_text)

    wb.close()
    return pages
